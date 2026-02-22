from fastapi import FastAPI, APIRouter, HTTPException, Body, BackgroundTasks
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta, time
import bcrypt
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Models
class Student(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    full_name: str
    class_grade: int  # 5-11
    class_subsection: str  # А, Б, В, Г, Д, Е, Ё
    shift: str  # "morning" or "afternoon" - now manually selectable
    face_descriptor: List[float]
    registered_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StudentCreate(BaseModel):
    full_name: str
    class_grade: int
    class_subsection: str
    shift: str  # Now required from frontend
    face_descriptor: List[float]

class Teacher(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    full_name: str
    subject: str  # Subject they teach (e.g., Mathematics, Informatics, English)
    face_descriptor: List[float]
    registered_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TeacherCreate(BaseModel):
    full_name: str
    subject: str
    face_descriptor: List[float]

class AttendanceRecord(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    student_id: str
    student_name: str
    class_grade: int
    class_subsection: str
    date: str  # YYYY-MM-DD
    scan_time: datetime
    shift: str
    status: str  # "present", "late", "absent"
    minutes_late: int = 0

class AttendanceScan(BaseModel):
    student_id: str
    face_descriptor: List[float]
    scan_time: Optional[str] = None

class DirectorLogin(BaseModel):
    username: str
    password: str

class Director(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    password_hash: str

# Helper Functions
def get_shift_start_time(shift: str) -> time:
    """Get shift start time"""
    if shift == "morning":
        return time(8, 0)
    else:  # afternoon
        return time(13, 0)

def calculate_lateness(scan_time: datetime, shift: str) -> tuple[str, int]:
    """Calculate if student is late and by how many minutes"""
    shift_start = get_shift_start_time(shift)
    scan_time_local = scan_time.replace(tzinfo=None)
    
    # Create datetime for shift start on the same day
    shift_start_dt = datetime.combine(scan_time_local.date(), shift_start)
    
    if scan_time_local <= shift_start_dt:
        return "present", 0
    else:
        minutes_late = int((scan_time_local - shift_start_dt).total_seconds() / 60)
        return "late", minutes_late

def is_before_shift_start(scan_time: datetime, shift: str) -> tuple[bool, str]:
    """Check if scan is before shift start time"""
    shift_start = get_shift_start_time(shift)
    scan_time_local = scan_time.replace(tzinfo=None)
    shift_start_dt = datetime.combine(scan_time_local.date(), shift_start)
    
    if scan_time_local < shift_start_dt:
        time_until = shift_start_dt - scan_time_local
        hours = int(time_until.total_seconds() // 3600)
        minutes = int((time_until.total_seconds() % 3600) // 60)
        return True, f"{shift_start.strftime('%H:%M')}"
    return False, ""

def euclidean_distance(desc1: List[float], desc2: List[float]) -> float:
    """Calculate Euclidean distance between two face descriptors"""
    if len(desc1) != len(desc2):
        return float('inf')
    return sum((a - b) ** 2 for a, b in zip(desc1, desc2)) ** 0.5

# Auto-absence background task
async def mark_auto_absences():
    """Mark students as absent if they haven't scanned 1 hour after shift start"""
    logger.info("Running auto-absence task...")
    
    now = datetime.now(timezone.utc)
    today_str = now.strftime("%Y-%m-%d")
    current_time = now.time()
    
    # Check morning shift (if after 09:00)
    if current_time >= time(9, 0):
        await process_absences_for_shift("morning", today_str)
    
    # Check afternoon shift (if after 14:00)
    if current_time >= time(14, 0):
        await process_absences_for_shift("afternoon", today_str)

async def process_absences_for_shift(shift: str, date_str: str):
    """Process absences for a specific shift"""
    # Get all students in this shift (exclude face_descriptor to save memory)
    students = await db.students.find({"shift": shift}, {"_id": 0, "face_descriptor": 0}).to_list(5000)
    
    # Get all attendance records for today
    attended = await db.attendance.find({"date": date_str, "shift": shift}, {"_id": 0}).to_list(5000)
    attended_ids = {record["student_id"] for record in attended}
    
    # Mark absent students
    for student in students:
        if student["id"] not in attended_ids:
            # Check if already marked absent
            existing = await db.attendance.find_one(
                {"student_id": student["id"], "date": date_str, "status": "absent"},
                {"_id": 0}
            )
            
            if not existing:
                absent_record = AttendanceRecord(
                    student_id=student["id"],
                    student_name=student["full_name"],
                    class_grade=student["class_grade"],
                    class_subsection=student["class_subsection"],
                    date=date_str,
                    scan_time=datetime.now(timezone.utc),
                    shift=shift,
                    status="absent",
                    minutes_late=0
                )
                
                doc = absent_record.model_dump()
                doc['scan_time'] = doc['scan_time'].isoformat()
                await db.attendance.insert_one(doc)
                logger.info(f"Marked {student['full_name']} as absent")

# Routes
@api_router.get("/")
async def root():
    return {"message": "Rejap Khujamov School Attendance System"}

# Director routes
@api_router.post("/auth/director/login")
async def director_login(credentials: DirectorLogin):
    director = await db.directors.find_one({"username": credentials.username}, {"_id": 0})
    
    if not director:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Verify password
    if bcrypt.checkpw(credentials.password.encode('utf-8'), director['password_hash'].encode('utf-8')):
        return {"success": True, "username": credentials.username}
    else:
        raise HTTPException(status_code=401, detail="Invalid credentials")

@api_router.post("/auth/director/create")
async def create_director(credentials: DirectorLogin):
    """Create a director account - for initial setup only"""
    # Check if director already exists
    existing = await db.directors.find_one({"username": credentials.username})
    if existing:
        raise HTTPException(status_code=400, detail="Director already exists")
    
    # Hash password
    password_hash = bcrypt.hashpw(credentials.password.encode('utf-8'), bcrypt.gensalt())
    
    director = Director(
        username=credentials.username,
        password_hash=password_hash.decode('utf-8')
    )
    
    doc = director.model_dump()
    await db.directors.insert_one(doc)
    
    return {"success": True, "message": "Director account created"}

# Student routes
@api_router.post("/students", response_model=Student)
async def register_student(student_data: StudentCreate):
    """Register a new student with face descriptor"""
    student = Student(
        full_name=student_data.full_name,
        class_grade=student_data.class_grade,
        class_subsection=student_data.class_subsection,
        shift=student_data.shift,  # Now from user input
        face_descriptor=student_data.face_descriptor
    )
    
    doc = student.model_dump()
    doc['registered_at'] = doc['registered_at'].isoformat()
    
    await db.students.insert_one(doc)
    return student

@api_router.get("/students", response_model=List[Student])
async def get_students():
    students = await db.students.find({}, {"_id": 0}).to_list(2000)
    
    for student in students:
        if isinstance(student.get('registered_at'), str):
            student['registered_at'] = datetime.fromisoformat(student['registered_at'])
    
    return students

@api_router.get("/students/class/{grade}/{subsection}")
async def get_students_by_class(grade: int, subsection: str):
    students = await db.students.find(
        {"class_grade": grade, "class_subsection": subsection},
        {"_id": 0, "face_descriptor": 0}
    ).to_list(100)
    return students

@api_router.delete("/students/{student_id}")
async def delete_student(student_id: str):
    """Delete a student and all their attendance records"""
    # Check if student exists
    student = await db.students.find_one({"id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    # Delete student
    await db.students.delete_one({"id": student_id})
    
    # Delete all attendance records
    await db.attendance.delete_many({"student_id": student_id})
    
    return {
        "success": True,
        "message": f"Student {student['full_name']} and all attendance records deleted"
    }

@api_router.get("/students/registration-stats")
async def get_registration_stats():
    """Get registration statistics - total students and registered students per class"""
    students = await db.students.find({}, {"_id": 0, "face_descriptor": 0}).to_list(2000)
    
    # Group by class
    stats = {}
    for student in students:
        class_key = f"{student['class_grade']}-{student['class_subsection']}"
        if class_key not in stats:
            stats[class_key] = {
                "class_name": class_key,
                "grade": student['class_grade'],
                "subsection": student['class_subsection'],
                "shift": student['shift'],
                "registered_count": 0,
                "students": []
            }
        stats[class_key]['registered_count'] += 1
        stats[class_key]['students'].append({
            "id": student['id'],
            "full_name": student['full_name'],
            "registered_at": student.get('registered_at', '')
        })
    
    return {
        "total_registered": len(students),
        "classes": list(stats.values())
    }

# Attendance routes
@api_router.post("/attendance/scan")
async def scan_attendance(scan_data: AttendanceScan):
    """Record attendance scan"""
    # Get scan time
    if scan_data.scan_time:
        scan_time = datetime.fromisoformat(scan_data.scan_time)
    else:
        scan_time = datetime.now(timezone.utc)
    
    today_str = scan_time.strftime("%Y-%m-%d")
    
    # Get student
    student = await db.students.find_one({"id": scan_data.student_id}, {"_id": 0})
    
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    # Verify face match (basic threshold check)
    distance = euclidean_distance(scan_data.face_descriptor, student['face_descriptor'])
    
    if distance > 0.6:  # Strict threshold
        raise HTTPException(status_code=403, detail="Face does not match")
    
    # Check if scanning before shift starts
    is_early, shift_start_time = is_before_shift_start(scan_time, student['shift'])
    if is_early:
        raise HTTPException(
            status_code=400,
            detail=f"too_early|{shift_start_time}"
        )
    
    # Check if already scanned today
    existing = await db.attendance.find_one(
        {"student_id": scan_data.student_id, "date": today_str},
        {"_id": 0}
    )
    
    if existing:
        raise HTTPException(status_code=400, detail="Already scanned today")
    
    # Calculate lateness
    status, minutes_late = calculate_lateness(scan_time, student['shift'])
    
    # Create attendance record
    record = AttendanceRecord(
        student_id=student['id'],
        student_name=student['full_name'],
        class_grade=student['class_grade'],
        class_subsection=student['class_subsection'],
        date=today_str,
        scan_time=scan_time,
        shift=student['shift'],
        status=status,
        minutes_late=minutes_late
    )
    
    doc = record.model_dump()
    doc['scan_time'] = doc['scan_time'].isoformat()
    
    await db.attendance.insert_one(doc)
    
    return {
        "success": True,
        "student_name": student['full_name'],
        "class": f"{student['class_grade']}-{student['class_subsection']}",
        "status": status,
        "minutes_late": minutes_late
    }

@api_router.get("/attendance/today")
async def get_today_attendance():
    """Get all attendance for today"""
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    records = await db.attendance.find({"date": today_str}, {"_id": 0}).to_list(2000)
    return records

@api_router.get("/attendance/class/{grade}/{subsection}")
async def get_class_attendance(grade: int, subsection: str):
    """Get attendance for a specific class"""
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # Get all students in class
    students = await db.students.find(
        {"class_grade": grade, "class_subsection": subsection},
        {"_id": 0, "face_descriptor": 0}
    ).to_list(100)
    
    # Get attendance records
    records = await db.attendance.find(
        {"class_grade": grade, "class_subsection": subsection, "date": today_str},
        {"_id": 0}
    ).to_list(100)
    
    return {
        "total_students": len(students),
        "attendance_records": records,
        "students": students
    }

@api_router.get("/analytics/weekly")
async def get_weekly_analytics():
    """Get weekly analytics for all classes"""
    # Get current week date range
    now = datetime.now(timezone.utc)
    week_start = now - timedelta(days=now.weekday())
    week_start_str = week_start.strftime("%Y-%m-%d")
    
    # Get all students grouped by class
    students = await db.students.find({}, {"_id": 0, "face_descriptor": 0}).to_list(2000)
    
    # Group by class
    classes = {}
    for student in students:
        class_key = f"{student['class_grade']}-{student['class_subsection']}"
        if class_key not in classes:
            classes[class_key] = {
                "class_name": class_key,
                "grade": student['class_grade'],
                "subsection": student['class_subsection'],
                "total_students": 0,
                "total_present": 0,
                "total_late": 0,
                "total_absent": 0,
                "total_lateness_minutes": 0,
                "attendance_rate": 0
            }
        classes[class_key]['total_students'] += 1
    
    # Get attendance records for this week
    records = await db.attendance.find(
        {"date": {"$gte": week_start_str}},
        {"_id": 0}
    ).to_list(5000)
    
    # Aggregate statistics
    for record in records:
        class_key = f"{record['class_grade']}-{record['class_subsection']}"
        if class_key in classes:
            if record['status'] == 'present':
                classes[class_key]['total_present'] += 1
            elif record['status'] == 'late':
                classes[class_key]['total_late'] += 1
                classes[class_key]['total_lateness_minutes'] += record['minutes_late']
            elif record['status'] == 'absent':
                classes[class_key]['total_absent'] += 1
    
    # Calculate attendance rates
    for class_data in classes.values():
        total_attendance = class_data['total_present'] + class_data['total_late']
        # Calculate over 5 school days
        total_possible = class_data['total_students'] * 5
        if total_possible > 0:
            class_data['attendance_rate'] = round((total_attendance / total_possible) * 100, 2)
    
    return list(classes.values())

@api_router.get("/analytics/class-of-week")
async def get_class_of_week():
    """Get the winning class of the week"""
    analytics = await get_weekly_analytics()
    
    if not analytics:
        return None
    
    # Find class with highest attendance and lowest lateness
    best_class = max(analytics, key=lambda x: (x['attendance_rate'], -x['total_lateness_minutes']))
    
    return best_class

@api_router.get("/export/weekly")
async def export_weekly_report():
    """Export weekly attendance report as Excel"""
    analytics = await get_weekly_analytics()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Weekly Attendance"
    
    # Header styling
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ["Class", "Total Students", "Present", "Late", "Absent", "Total Lateness (min)", "Attendance Rate (%)"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for row, class_data in enumerate(analytics, 2):
        ws.cell(row=row, column=1, value=class_data['class_name'])
        ws.cell(row=row, column=2, value=class_data['total_students'])
        ws.cell(row=row, column=3, value=class_data['total_present'])
        ws.cell(row=row, column=4, value=class_data['total_late'])
        ws.cell(row=row, column=5, value=class_data['total_absent'])
        ws.cell(row=row, column=6, value=class_data['total_lateness_minutes'])
        ws.cell(row=row, column=7, value=class_data['attendance_rate'])
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=weekly_attendance_report.xlsx"}
    )

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Schedule auto-absence task
scheduler = BackgroundScheduler()
scheduler.add_job(
    lambda: app.state.loop.create_task(mark_auto_absences()),
    CronTrigger(hour='9,14', minute=5),  # Run at 09:05 and 14:05
    id='auto_absence_task'
)

@app.on_event("startup")
async def startup_event():
    import asyncio
    app.state.loop = asyncio.get_event_loop()
    scheduler.start()
    logger.info("Auto-absence scheduler started")

@app.on_event("shutdown")
async def shutdown_db_client():
    scheduler.shutdown()
    client.close()
